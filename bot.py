import os
import asyncio
import logging
import requests
import sqlite3
import html
import re
import pandas as pd
from io import BytesIO
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from bs4 import BeautifulSoup
from aiohttp import web
from datetime import datetime, timedelta
from telethon import TelegramClient, events
from telethon.sessions import StringSession
from openpyxl.styles import Font, PatternFill, Alignment
from telethon.errors import FloodWaitError
from aiogram.utils.exceptions import MessageNotModified
from contextlib import suppress
from curl_cffi import requests as crequests

# --- НАСТРОЙКИ ---
TOKEN = os.getenv('BOT_TOKEN')
API_ID = 23009673
API_HASH = '249328ef42a91e5c80102c3d73c76a9c'
SESSION_STR = os.getenv('TELEGRAM_SESSION')
SJ_KEY = os.getenv('SUPERJOB_KEY') # Твой новый ключ
HH_TOKEN = os.getenv('HH_TOKEN')
# Список каналов БЕЗ собаки @
CHANNELS = [
    'vdhl_good', 'mediajobs_ru', 'work_in_media', 'distantsiya',
    'theblueprintcareer', 'huggabletalents', 'careerspace', 'morejobs', 'heyanie', 
    'marketing_jobs', 'mirkreatorovjob', 'budujobs', 'normrabota', 'jobpower', 
    'it_vakansii_jobs', 'workasap', 'forproducer', 'vacanciesrus', 
    'honeyiwantmoney', 'evacuatejobs', 'workinart', 'young_relocate', 'iRecommendWork_IT',
    'Inwork', 'cliquejobs', 'fashionfaculty', 'forallmedia', 'vitrinajobs',
    'dnative_job', 'Young_and_Yandex', 'digital_rabota', 'jobforpr', 'megafonjobs',
    'cozy_hr', 'rassvet_pro', 'young_june', 'mtsfintechjobs', 'digital_hr',
    'zdemcv', 'foranalysts', 'hcareers_jobs', 'product_jobs', 'edujobs',
    'talentswanted', 'forchiefs', 'rabota_marketing_juniors', 'tj_collega', 'remotegeekjob',
    'forallmarketing', 'hsecareer', 'dddwork', 'ya_jobs_pm', 'forproducts',
    'practicum_experts', 'zloy_kollega', 'jobforjunior', 'vacantcist', 'avito_career',
    'workasap', 'mirkreatorovjob', 'careerspace', 'huggabletalents', 'theblueprintcareer',
    'heyanie', 'marketing_jobs', 'it_vakansii_jobs'
]

# Создаем клиента для чтения каналов
client = TelegramClient(StringSession(SESSION_STR), API_ID, API_HASH)


bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
logging.basicConfig(level=logging.INFO)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Cache-Control': 'no-cache',
    'Pragma': 'no-cache',
    'Sec-Ch-Ua': '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1'
}

# --- БАЗА ДАННЫХ ---
def init_db():
    conn = sqlite3.connect('manager.db')
    conn.execute('CREATE TABLE IF NOT EXISTS jobs (id TEXT PRIMARY KEY)')
    conn.execute('CREATE TABLE IF NOT EXISTS subs (user_id INTEGER, keyword TEXT, UNIQUE(user_id, keyword))')
    conn.commit()
    conn.close()

def add_subscription(user_id, keyword):
    conn = sqlite3.connect('manager.db')
    conn.execute('INSERT OR IGNORE INTO subs (user_id, keyword) VALUES (?, ?)', (user_id, keyword.lower()))
    conn.commit()
    conn.close()

def get_all_subs():
    conn = sqlite3.connect('manager.db')
    data = conn.execute('SELECT user_id, keyword FROM subs').fetchall()
    conn.close()
    return data

def is_new_job(job_id):
    conn = sqlite3.connect('manager.db')
    res = conn.execute('SELECT 1 FROM jobs WHERE id = ?', (job_id,)).fetchone()
    if not res:
        conn.execute('INSERT INTO jobs VALUES (?)', (job_id,))
        conn.commit()
        conn.close()
        return True
    conn.close()
    return False

# --- ПАРСЕРЫ ---
async def search_telegram_history(query, limit_per_channel=2):
    if not client.is_connected(): 
        await client.start()
        
    results = []
    seen_texts = set()
    # Ищем вакансии не старше 14 дней
    date_limit = datetime.now() - timedelta(days=14)
    query_low = query.lower().replace('ё', 'е')

    # Проходим по списку (лучше ограничить до 30 для скорости поиска)
    for channel in CHANNELS[:30]:
        try:
            # search=query ищет на серверах ТГ, limit=2 берет только самые свежие
            async for msg in client.iter_messages(channel, search=query, limit=limit_per_channel):
                
                # 1. Фильтр на мусор (короткие сообщения) и дату
                if not msg.text or len(msg.text) < 100: continue
                if msg.date.replace(tzinfo=None) < date_limit: continue

                # 2. Защита от дублей (репосты в разных каналах)
                text_id = msg.text[:100].lower().replace('ё', 'е').strip()
                if text_id in seen_texts: continue
                seen_texts.add(text_id)

                # 3. Поиск зарплаты (улучшенная регулярка)
                pay = "См. в посте"
                salary_found = re.search(r'(\d[\d\s\.]*)\s?(руб|р\.|₽|\$|€|usd|eur|к|k)', msg.text.lower())
                if salary_found:
                    pay = salary_found.group(0).strip()
                    # --- ВСТАВИТЬ ЭТО ---
                # Извлекаем начало поста как суть вакансии
                # Убираем переносы строк, чтобы текст был плотным
                desc = msg.text[:200].replace('\n', ' ').replace('*', '').replace('_', '').strip()
                # --------------------

                # 4. ТВОЙ ЛЮБИМЫЙ ФОРМАТ ВЫВОДА
                # Очищаем текст от символов, которые могут сломать Markdown
                display_text = msg.text[:400].replace('*', '').replace('_', '').strip()
                
                results.append({
                    'id': f"tg_{channel}_{msg.id}",
                    # Строгий формат: Иконка - Канал - Текст - ЗП (если нашли) - Ссылка
                    'text': f"📱 TG [{channel}]: {display_text}...\n\n💰 Зарплата: {pay}\nhttps://t.me/{channel}/{msg.id}",
                    'Дата': msg.date.strftime('%Y-%m-%d'),
                    'Источник': f'TG: {channel}',
                    'Вакансия': query.capitalize(),
                    'Компания': channel,
                    'Оплата': pay,
                    'Ссылка': f"https://t.me/{channel}/{msg.id}",
                    'Описание': desc
                })
            
            # Микро-пауза для обхода FloodWait
            await asyncio.sleep(0.3) 
            
        except Exception as e:
            logging.error(f"Ошибка канала {channel}: {e}")
            continue
    return results

def search_trudvsem(query, limit=50):
    # Официальный API "Работа России" (Москва)
    url = f"https://opendata.trudvsem.ru/api/v1/vacancies/region/77?text={query}&limit={limit}"
    results = []
    try:
        r = requests.get(url, timeout=15)
        if r.status_code != 200: return []
        
        data = r.json()
        # Трудвсем отдает вложенный список
        vacancies = data.get('results', {}).get('vacancies', [])
        
        for item in vacancies:
            v = item.get('vacancy', {})
            
            # Чистим требования от HTML мусора
            raw_req = v.get('requirement', 'Описание доступно по ссылке')
            clean_req = re.sub(r'<[^>]*>', '', str(raw_req)).replace('&quot;', '"').strip()

            results.append({
                'id': f"trud_{v.get('id')}",
                'Дата': v.get('creation-date', '')[:10],
                'Источник': 'Trud', 
                'Вакансия': v.get('job-name', '—'), 
                'Компания': v.get('company', {}).get('name', '—'), 
                'Оплата': v.get('salary', 'Договорная'), 
                'Ссылка': v.get('vac_url', '#'),
                'Описание': clean_req[:250]
            })
            
        logging.info(f"TrudVsem: Найдено {len(results)} вакансий")
    except Exception as e:
        logging.error(f"TrudVsem error: {e}")
    
    return results

def search_hh(query, limit=100):
    GOOGLE_PROXY_URL = "https://script.google.com/macros/s/AKfycbz89VYCumV1LC4-52i33YYdoFO5MCfCMwZE_ZR6SagJc73enQuXng8mq37zsougaj1TPA/exec"
    results = []
    try:
        r = requests.get(f"{GOOGLE_PROXY_URL}?q={query}", timeout=30)
        if r.status_code != 200: return []
        
        soup = BeautifulSoup(r.text, 'html.parser')
        # В твоем логе вакансии лежат в div-ах с классом serp-item
        items = soup.find_all('div', class_='serp-item')
        
        for v in items:
            try:
                # Ищем заголовок внутри data-qa
                title_el = v.find('a', {'data-qa': 'serp-item__title'})
                if not title_el: continue
                
                title = title_el.text.strip()
                # Получаем чистую ссылку
                link = title_el['href'].split('?')[0]
                if not link.startswith('http'): link = 'https://hh.ru' + link
                
                # Зарплата
                salary_el = v.find('span', {'data-qa': 'vacancy-serp__vacancy-compensation'})
                pay = salary_el.text.strip() if salary_el else "Договорная"
                
                # Компания
                comp_el = v.find('a', {'data-qa': 'vacancy-serp__vacancy-employer'})
                company = comp_el.text.strip() if comp_el else "—"

                # Суть (Snippet)
                desc_el = v.find('div', {'data-qa': 'vacancy-serp__vacancy_snippet_requirement'})
                desc = desc_el.text.strip() if desc_el else "Описание в вакансии"

                results.append({
                    'id': f"hh_{hash(link)}",
                    'Дата': datetime.now().strftime('%Y-%m-%d'),
                    'Источник': 'HH', 
                    'Вакансия': title, 
                    'Компания': company, 
                    'Оплата': pay, 
                    'Ссылка': link,
                    'Описание': desc[:250]
                })
            except: continue
        logging.info(f"HH: Найдено {len(results)} вакансий")
    except Exception as e:
        logging.error(f"HH Error: {e}")
    return results
    
def search_superjob(query, limit=50):
    if not SJ_KEY: return []
    url = f"https://api.superjob.ru/2.0/vacancies/?keyword={query}&town=4&count={limit}"
    headers = {'X-Api-App-Id': SJ_KEY}
    results = []
    try:
        r = requests.get(url, headers=headers, timeout=10).json()
        for v in r.get('objects', []):
            # ВНИМАНИЕ: Обязательно вычисляем pay перед использованием!
            p_from = v.get('payment_from')
            p_to = v.get('payment_to')
            if p_from and p_to: pay = f"{p_from}-{p_to}"
            elif p_from: pay = f"от {p_from}"
            else: pay = "Договорная"
                # --- ВСТАВИТЬ ЭТО ---
            # Извлекаем краткое описание (суть) из поля candidat
            desc = (v.get('candidat', '') or "").replace('\n', ' ').strip()
            # --------------------

            results.append({
                'id': f"sj_{v['id']}",
                'text': f"🔵 SJ: {v['profession']}\n💰 {pay} | 📅 {datetime.fromtimestamp(v['date_published']).strftime('%d.%m')}\n{v['link']}",
                'Дата': datetime.fromtimestamp(v['date_published']).strftime('%Y-%m-%d'),
                'Источник': 'SuperJob', 'Вакансия': v['profession'], 'Компания': v['client'].get('title', '—'), 'Оплата': pay, 'Ссылка': v['link'],
                'Описание': desc[:200]
            })
    except Exception as e:
        logging.error(f"SJ error: {e}")
    return results
    
def search_habr(query, limit=20):
    # Москва = 678, тип вакансий = все
    url = f"https://career.habr.com/vacancies?q={query}&city_id=678&type=all"
    results = []
    try:
        # Устанавливаем таймаут чуть больше, Хабр иногда думает
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            logging.error(f"Habr returned status {r.status_code}")
            return []

        soup = BeautifulSoup(r.text, 'html.parser')

        # Основной контейнер вакансий на Хабре
        items = soup.find_all('div', class_='vacancy-card')

        for i in items[:limit]:
            # Ищем заголовок и ссылку
            title_link = i.find('a', class_='vacancy-card__title-link')
            company_link = i.find('a', class_='vacancy-card__company-title') # Ссылка на компанию
            salary_div = i.find('div', class_='basic-salary') # Зарплата

            if title_link:
                title = title_link.text.strip()
                link = "https://career.habr.com" + title_link['href']
                company = company_link.text.strip() if company_link else "Компания не указана"
                pay = salary_div.text.strip() if salary_div else "Договорная"
                # --- ВСТАВИТЬ ЭТО ---
                # Ищем навыки (стек технологий) как суть вакансии
                skills_div = i.find('div', class_='vacancy-card__skills')
                desc = skills_div.text.strip() if skills_div else "Стек технологий доступен по ссылке."
                # --------------------

                # Сохраняем в нашем эталонном формате
                results.append({
                    'id': f"hb_{link.split('/')[-1]}", # Берем ID вакансии из URL
                    'text': f"🟢 Habr: {title}\n💰 {pay} | 📅 Сегодня\n{link}",
                    'Дата': datetime.now().strftime('%Y-%m-%d'), # Хабр пишет "вчера/сегодня", для Excel ставим текущую
                    'Источник': 'Habr',
                    'Вакансия': title,
                    'Компания': company,
                    'Оплата': pay,
                    'Ссылка': link,
                    'Описание': desc
                })

        logging.info(f"Habr: Найдено {len(results)} вакансий")

    except Exception as e:
        logging.error(f"Критическая ошибка Хабр Карьеры: {e}")

    return results

def search_geekjob(query, limit=10):
    # Поиск по разделу Digital/IT
    url = f"https://geekjob.ru/vacancies?q={query.replace(' ', '+')}"
    results = []
    try:
        r = crequests.get(url, headers=HEADERS, impersonate="chrome120", timeout=10)
        soup = BeautifulSoup(r.text, 'html.parser')
        
        # Ищем карточки вакансий
        items = soup.find_all('li', class_='vacancy-card')
        
        for i in items[:limit]:
            a = i.find('a', class_='vacancy-name')
            if a:
                title = a.text.strip()
                link = "https://geekjob.ru" + a['href']
                
                # Ищем компанию
                comp_div = i.find('div', class_='company-name')
                company = comp_div.text.strip() if comp_div else "—"
                
                # Ищем зарплату
                pay_span = i.find('span', class_='salary')
                pay = pay_span.text.strip() if pay_span else "Договорная"
                
                # --- НОВОЕ: ИЩЕМ СУТЬ (ОПИСАНИЕ) ---
                # На GeekJob суть обычно в блоке vacancy-description или в тегах
                desc_div = i.find('div', class_='vacancy-description') or i.find('p')
                desc = desc_div.text.strip() if desc_div else "IT-вакансия: подробности стека и задач доступны по ссылке."

                results.append({
                    'id': f"gj_{link.split('/')[-1]}",
                    'Дата': datetime.now().strftime('%Y-%m-%d'),
                    'Источник': 'GeekJob', 
                    'Вакансия': title, 
                    'Компания': company, 
                    'Оплата': pay, 
                    'Ссылка': link,
                    'Описание': desc[:200] # Наша "суть" для ручного поиска
                })
    except Exception as e:
        logging.error(f"GeekJob error: {e}")
        
    return results

def search_jobfilter(query, limit=5):
    url = f"https://jobfilter.ru/vacancies?q={query.replace(' ', '+')}&city=москва"
    results = []
    try:
        r = crequests.get(url, headers=HEADERS, impersonate="chrome120", timeout=10)
        soup = BeautifulSoup(r.text, 'html.parser')
        
        # Находим карточки вакансий
        items = soup.find_all('div', class_='vacancy_item') or soup.find_all('div', class_='vacancy-item')
        
        for i in items[:limit]:
            a = i.find('a')
            if a:
                title = a.text.strip()
                link = "https://jobfilter.ru" + a['href'] if not a['href'].startswith('http') else a['href']
                
                # --- ИЩЕМ ОПИСАНИЕ (СУТЬ) ---
                desc_div = i.find('div', class_='description') or i.find('div', class_='vacancy_description')
                desc = desc_div.text.strip() if desc_div else "Краткое описание доступно по ссылке на сайте."
                
                # --- ИЩЕМ ЗАРПЛАТУ (если она есть в карточке) ---
                pay_div = i.find('div', class_='vacancy_salary') or i.find('span', class_='salary')
                pay = pay_div.text.strip() if pay_div else "См. на сайте"
                
                # --- ИЩЕМ КОМПАНИЮ ---
                comp_div = i.find('div', class_='company') or i.find('span', class_='company')
                company = comp_div.text.strip() if comp_div else "Компания не указана"

                results.append({
                    'id': f"jf_{hash(link)}", # Используем хэш ссылки как уникальный ID
                    'Дата': datetime.now().strftime('%Y-%m-%d'),
                    'Источник': 'JobFilter',
                    'Вакансия': title,
                    'Компания': company,
                    'Оплата': pay,
                    'Ссылка': link,
                    'Описание': desc[:200] # Добавляем ключ Описание (суть)
                })
    except Exception as e:
        logging.error(f"JobFilter error: {e}")
        
    return results

@client.on(events.NewMessage(chats=CHANNELS))
async def telethon_handler(event):
    try:
        # 1. Получаем текст и приводим к нижнему регистру + убираем ё
        text = event.message.message
        if not text: return
        text_for_search = text.lower().replace('ё', 'е')

        # 2. Получаем данные о канале
        chat = await event.get_chat()
        title = getattr(chat, 'title', 'Media Channel')
        username = getattr(chat, 'username', 'channel')

        # 3. Проверяем подписки
        subs = get_all_subs()
        matched_users = []

        for user_id, kw in subs:
            # Сравниваем ключевое слово (тоже без ё) с текстом поста
            if kw.lower().replace('ё', 'е') in text_for_search:
                matched_users.append(user_id)

        if matched_users:
            # 4. Проверка на дубликаты (чтобы не слать одно и то же)
            if is_new_job(f"tg_{event.chat_id}_{event.id}"):
                # Формируем прямую ссылку на пост
                post_url = f"https://t.me/{username}/{event.id}" if username else "Ссылка скрыта"

                for uid in set(matched_users):
                    try:
                        # 5. Отправляем сообщение с заголовком и ссылкой в конце
                        await bot.send_message(
                            uid, 
                            f"⚡️ **НОВОЕ В КАНАЛЕ: {title}**\n\n"
                            f"{text[:3500]}\n\n"
                            f"🔗 **Оригинал поста:** {post_url}",
                            disable_web_page_preview=False # Оставляем превью для удобства
                        )
                        await asyncio.sleep(0.3) # Защита от Flood Limit
                    except: pass
    except Exception as e:
        logging.error(f"Ошибка в telethon_handler: {e}")

async def monitor_sites():
    while True:
        try:
            logging.info("Запуск циклической проверки сайтов...")
            subs = get_all_subs() # Получаем список [(user_id, keyword), ...]

            if not subs:
                logging.info("Подписок пока нет. Спим.")
                await asyncio.sleep(600)
                continue

            for user_id, kw in subs:
                # 1. Собираем свежак с 3-х главных сайтов (по 15 штук)
                # Habr и JobFilter тоже можно добавить, если нужно
                hh = search_hh(kw, 15)
                sj = search_superjob(kw, 15)
                hb = search_habr(kw, 10) 

                all_current_jobs = hh + sj + hb

                # Нормализуем ключевое слово для проверки (е/ё)
                kw_clean = kw.lower().replace('ё', 'е')

                for job in all_current_jobs:
                    # 2. Проверяем, подходит ли вакансия под ключевое слово (на всякий случай)
                    # и не присылали ли мы её уже раньше (is_new_job)
                    job_text_clean = job['text'].lower().replace('ё', 'е')

                    if kw_clean in job_text_clean:
                        if is_new_job(job['id']):
                            try:
                                # 3. Отправляем пользователю
                                await bot.send_message(
                                    user_id, 
                                    f"🔔 **НОВАЯ ВАКАНСИЯ ПО ПОДПИСКЕ**\n"
                                    f"🔍 Запрос: `{kw.upper()}`\n\n"
                                    f"{job['text']}"
                                )
                                # Небольшая пауза, чтобы не поймать Flood Limit от Telegram
                                await asyncio.sleep(0.7)
                            except Exception as send_error:
                                logging.error(f"Ошибка отправки сообщения юзеру {user_id}: {send_error}")

            logging.info("Проверка завершена. Следующий запуск через 30 минут.")

        except Exception as e:
            logging.error(f"Критическая ошибка в monitor_sites: {e}")

        # Ждем 30 минут (1800 секунд) перед следующей проверкой
        await asyncio.sleep(1800)

def generate_excel(data):
    try:
        if not data:
            return None
            
        raw_data = []
        for item in data:
            # 1. Добавляем 'Описание' в список собираемых полей
            raw_data.append({
                'Дата': str(item.get('Дата', '—')),
                'Источник': str(item.get('Источник', '—')),
                'Вакансия': str(item.get('Вакансия', '—')),
                'Компания': str(item.get('Компания', '—')),
                'Оплата': str(item.get('Оплата', '—')),
                'Описание': str(item.get('Описание', '—')), # НОВОЕ ПОЛЕ
                'Ссылка': str(item.get('Ссылка', '—'))
            })

        df = pd.DataFrame(raw_data)
        
        # 2. Умная сортировка по дате
        if not df.empty:
            df['TempDate'] = pd.to_datetime(df['Дата'], errors='coerce')
            df = df.sort_values(by='TempDate', ascending=False)
            df = df.drop(columns=['TempDate']) 

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Вакансии')
            ws = writer.sheets['Вакансии']
            
            # 3. Настройка оформления
            header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            link_font = Font(color="0000FF", underline="single")
            
            # Находим индексы нужных колонок
            link_col_idx = df.columns.get_loc("Ссылка") + 1 if "Ссылка" in df.columns else None
            desc_col_idx = df.columns.get_loc("Описание") + 1 if "Описание" in df.columns else None

            for col_num, column in enumerate(df.columns):
                cell = ws.cell(row=1, column=col_num + 1)
                cell.fill = header_fill
                cell.font = header_font
                
                col_letter = chr(65 + col_num)
                # Настройка ширины колонок
                if column == 'Ссылка':
                    ws.column_dimensions[col_letter].width = 45
                elif column == 'Вакансия':
                    ws.column_dimensions[col_letter].width = 40
                elif column == 'Описание':
                    ws.column_dimensions[col_letter].width = 60 # Широкая колонка для сути
                else:
                    ws.column_dimensions[col_letter].width = 20

            # 4. Делаем ссылки кликабельными и настраиваем перенос текста для Описания
            for row in range(2, len(df) + 2):
                # Для ссылок
                if link_col_idx:
                    cell_link = ws.cell(row=row, column=link_col_idx)
                    if cell_link.value and str(cell_link.value).startswith('http'):
                        cell_link.hyperlink = cell_link.value
                        cell_link.font = link_font
                
                # Для описания (включаем перенос текста, чтобы было удобно читать)
                if desc_col_idx:
                    cell_desc = ws.cell(row=row, column=desc_col_idx)
                    cell_desc.alignment = Alignment(wrap_text=True, vertical='top')

            ws.freeze_panes = 'A2'
            
        output.seek(0)
        return output
    except Exception as e:
        logging.error(f"Критическая ошибка Excel: {e}")
        return None

# --- ОБРАБОТЧИКИ ---

@dp.message_handler(commands=['start'])
async def start_cmd(message: types.Message):
    name = message.from_user.first_name
    await message.answer(
        f"Привет, {name}! 🎬 Я — твой персональный агент по поиску работы в кино и медиа.\n\n"
        f"**Что я умею:**\n"
        f"🔍 **Мгновенный поиск:** Напиши название профессии, и я тут же перерою HH.ru, SuperJob, Habr и JobFilter.\n"
        f"📂 **Excel-отчеты:** На каждый запрос я присылаю файл с 50 свежими вакансиями.\n"
        f"⚡️ **Live-мониторинг:** Я читаю Telegram-каналы в реальном времени.\n\n"
        f"**Как запустить авто-поиск:**\n"
        f"1️⃣ Напиши ключевое слово, например: `режиссер` или `продюсер`.\n"
        f"2️⃣ Под результатом поиска нажми кнопку **«🔔 Подписаться»**.\n"
        f"3️⃣ Всё! Как только в каналах или на сайтах появится вакансия с этим словом — я мгновенно пришлю её тебе в личку.\n\n"
        f"💡 **Совет:** Подписывайся на короткие слова (например, `режиссер`), чтобы я ловил все склонения: *«ищем режиссера»*, *«нужны режиссеры»*.\n\n"
        f"Что ищем сегодня?",
        parse_mode="Markdown"
    )

@dp.message_handler(commands=['mysubs'])
async def list_subs(message: types.Message):
    conn = None
    try:
        # Устанавливаем таймаут, чтобы подождать, если база занята
        conn = sqlite3.connect('manager.db', timeout=20)
        cursor = conn.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS subs (user_id INTEGER, keyword TEXT, UNIQUE(user_id, keyword))')
        
        data = cursor.execute('SELECT keyword FROM subs WHERE user_id = ?', (message.from_user.id,)).fetchall()
        
        if not data:
            await message.answer("У тебя пока нет активных подписок. 🤷‍♂️")
        else:
            # Используем HTML, чтобы не конфликтовать с Markdown
            subs_list = "\n".join([f"✅ <code>{html.escape(row[0])}</code>" for row in data])
            await message.answer(f"🔔 <b>Твои подписки:</b>\n\n{subs_list}\n\n🗑 <code>/del слово</code>", parse_mode="HTML")
            
    except Exception as e:
        logging.error(f"Ошибка в mysubs: {e}")
        await message.answer("⚠ Ошибка базы данных. Попробуй еще раз через 5 секунд.")
    finally:
        if conn:
            conn.close()
        
@dp.message_handler(commands=['del'])
async def del_sub(message: types.Message):
    # Извлекаем слово, которое идет после команды /del
    keyword = message.get_args().lower().strip()

    if not keyword:
        await message.answer(
            "⚠ Нужно указать слово для удаления.\n\nПример: `/del режиссер`", 
            parse_mode="Markdown"
        )
        return

    conn = sqlite3.connect('manager.db')
    # Проверяем, была ли такая подписка
    res = conn.execute('SELECT 1 FROM subs WHERE user_id = ? AND keyword = ?', (message.from_user.id, keyword)).fetchone()

    if res:
        conn.execute('DELETE FROM subs WHERE user_id = ? AND keyword = ?', (message.from_user.id, keyword))
        conn.commit()
        await message.answer(f"❌ Подписка на слово '{keyword}' удалена. Я больше не буду присылать по нему уведомления.")
    else:
        await message.answer(f"🤔 У тебя нет активной подписки на слово '{keyword}'. Проверь список командой /mysubs")
    conn.close()

@dp.message_handler(commands=['stop_all'])
async def clear_subs(message: types.Message):
    conn = sqlite3.connect('manager.db')
    conn.execute('DELETE FROM subs WHERE user_id = ?', (message.from_user.id,))
    conn.commit()
    conn.close()
    await message.answer("📴 Мониторинг остановлен. Все твои подписки были успешно удалены.")

@dp.message_handler()
async def manual_search(message: types.Message):
    if message.text.startswith('/'): return

    query = message.text.strip()
    # Ограничиваем запрос для кнопки (Telegram limit 64 bytes)
    short_query = query[:30] 
    
    status_msg = await message.answer(f"🔎 <b>Ищу вакансии:</b> <code>{query}</code>...", parse_mode="HTML")

    all_found = []
    seen_ids = set()

    # 1. Сбор с сайтов
    with suppress(MessageNotModified):
        await status_msg.edit_text("🌐 <b>Опрашиваю сайты...</b>", parse_mode="HTML")

    loop = asyncio.get_running_loop()
    # Добавляем JobFilter обратно, он давал массу результатов
    tasks = [
        loop.run_in_executor(None, search_hh, query, 50),
        loop.run_in_executor(None, search_superjob, query, 30),
        loop.run_in_executor(None, search_habr, query, 20),
        loop.run_in_executor(None, search_trudvsem, query, 40), # Прямой метод
        loop.run_in_executor(None, search_jobfilter, query, 15) 
    ]
    
    try:
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for res in results:
            if isinstance(res, list):
                for job in res:
                    if job.get('id') and job['id'] not in seen_ids:
                        all_found.append(job)
                        seen_ids.add(job['id'])
    except Exception as e:
        logging.error(f"Ошибка сайтов: {e}")

    # 2. Сбор из Telegram
    with suppress(MessageNotModified):
        # Показываем, сколько уже нашли на сайтах
        await status_msg.edit_text(f"📡 <b>Проверяю 54 канала...</b> (Сайты: {len(all_found)})", parse_mode="HTML")

    try:
        # Увеличим лимит до 5, чтобы не пропускать свежее
        tg_hist = await search_telegram_history(query, limit_per_channel=5)
        if tg_hist:
            for job in tg_hist:
                if job['id'] not in seen_ids:
                    all_found.append(job)
                    seen_ids.add(job['id'])
    except Exception as e:
        logging.error(f"Ошибка TG: {e}")

    if not all_found:
        await status_msg.edit_text(f"❌ По запросу '<b>{query}</b>' ничего не найдено.", parse_mode="HTML")
        return

    # 3. УМНАЯ СОРТИРОВКА
    def sort_key(x):
        d = str(x.get('Дата', '')).lower()
        # Если дата YYYY-MM-DD (как у Trud и HH), она сортируется идеально сама
        # Если дата DD.MM, превращаем ее для сравнения
        if '.' in d and len(d) <= 5:
            day, month = d.split('.')
            return f"2026-{month}-{day}"
        return d if d else "0000-00-00"

    all_found.sort(key=sort_key, reverse=True)

    # 4. Вывод топ-15
    for j in all_found[:15]:
        try:
            v_name = html.escape(str(j.get('Вакансия', '—')))
            v_pay = html.escape(str(j.get('Оплата', 'Договорная')))
            v_comp = html.escape(str(j.get('Компания', '—')))
            v_date = html.escape(str(j.get('Дата', '—')))
            v_src = str(j.get('Источник', ''))
            v_link = j.get('Ссылка', '#')
            # Берем наше новое поле Описание
            v_desc = html.escape(str(j.get('Описание', 'Описание доступно по ссылке.')))

            # Подбираем иконку и текстовую метку
            if "HH" in v_src: 
                icon, label = "🔴", "HH"
            elif "SJ" in v_src or "SuperJob" in v_src: 
                icon, label = "🔵", "SJ"
            elif "Habr" in v_src: 
                icon, label = "🟢", "Habr"
            elif "Trud" in v_src: 
                icon, label = "🏢", "Trud" # Исправлено: теперь две переменные
            elif "Geek" in v_src: 
                icon, label = "👨‍💻", "GeekJob"
            else: 
                icon, label = "📱", "TG"

            # ФОРМИРУЕМ КРАСИВЫЙ ТЕКСТ (Как в подписке)
            pretty_text = (
                f"{icon} {label}: <b>{v_name}</b>\n"
                f"💰 <b>{v_pay}</b> | 📅 {v_date}\n"
                f"🏢 {v_comp}\n\n"
                f"📝 <b>Суть:</b> <i>{v_desc}...</i>\n\n"
                f"{v_link}"
            )

            await message.answer(pretty_text, disable_web_page_preview=True, parse_mode="HTML")
            await asyncio.sleep(0.3)
        except Exception as e:
            logging.error(f"Ошибка вывода: {e}")

    # 5. Генерация Excel и кнопка
    try:
        excel_file = generate_excel(all_found)
        if excel_file:
            # Важно: используем short_query, чтобы кнопка не сломалась
            kb = types.InlineKeyboardMarkup().add(
                types.InlineKeyboardButton(f"🔔 Подписаться на '{short_query}'", callback_data=f"sub|{short_query}")
            )
            await message.answer_document(
                types.InputFile(excel_file, filename=f"Jobs_{short_query}.xlsx"),
                caption=f"✅ Найдено вакансий: <b>{len(all_found)}</b>\n(Сортировка: Самые новые вверху)",
                reply_markup=kb,
                parse_mode="HTML"
            )
    except Exception as e:
        logging.error(f"Excel error: {e}")
        await message.answer(f"📊 Поиск завершен. Найдено вакансий: {len(all_found)}")

    await status_msg.delete()

# --- ОБРАБОТЧИК КНОПКИ ПОДПИСКИ ---
@dp.callback_query_handler(lambda c: c.data.startswith('sub|'))
async def sub_handler(cb: types.CallbackQuery):
    # 1. Извлекаем ключевое слово из даты кнопки
    kw = cb.data.split('|')[1]
    user_id = cb.from_user.id

    try:
        # 2. Записываем в базу данных
        add_subscription(user_id, kw)

        # 3. Отправляем всплывающее уведомление (alert)
        await bot.answer_callback_query(
            cb.id, 
            text=f"✅ Подписка на '{kw}' оформлена!", 
            show_alert=True
        )

        # 4. Отправляем подтверждающее сообщение в чат
        await bot.send_message(
            user_id, 
            f"🔔 **Готово!**\n\nЯ запомнил запрос: `{kw}`.\n"
            f"Как только в каналах или на сайтах появится новая вакансия с этим словом, я мгновенно пришлю её тебе сюда."
        )

    except Exception as e:
        logging.error(f"Subscription error: {e}")
        await bot.answer_callback_query(cb.id, text="❌ Произошла ошибка при подписке.")

async def handle(request):
    return web.Response(text="Bot is running!")

async def main(): # НИКАКИХ ОТСТУПОВ ПЕРЕД async
    init_db()

    # 1. Запуск Веб-сервера
    app = web.Application()
    app.router.add_get('/', handle)
    runner = web.AppRunner(app)
    await runner.setup()
    await web.TCPSite(runner, '0.0.0.0', int(os.environ.get("PORT", 8080))).start()

    # 2. Запуск мониторинга каналов
    try:
        await client.connect() # Подключаемся
        if not await client.is_user_authorized():
            logging.error("Telethon не авторизован!")
            # Если ты на Render, тут ничего не поделать, нужна новая StringSession
        
        # МАГИЯ: Запрашиваем диалоги, чтобы сбросить Timestamp
        await client.get_dialogs(limit=1)
        logging.info("Telethon синхронизирован!")
    except Exception as e:
        logging.error(f"Telethon error: {e}")

    # 3. Бот
    # Если функции monitor_sites нет, закомментируй строку ниже:
    asyncio.create_task(monitor_sites()) 
    await dp.start_polling()

if __name__ == '__main__':
    asyncio.run(main())
